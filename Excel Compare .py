#!/usr/bin/env python
# coding: utf-8

# In[6]:


import openpyxl
from openpyxl.styles import PatternFill, colors


# In[7]:


#Load in file and sheet
file1 = openpyxl.load_workbook("Oldfile.xlsx")
file2 = openpyxl.load_workbook("Newfile.xlsx")

sheet1 = file1['Sheet1']
sheet2 = file2['Sheet1']


# In[8]:


#Check if load was sucess 
print(sheet2['A2'].value)


# In[9]:


#Define the color to fill

#Orange
fill_style = PatternFill(start_color = 'FFA500', end_color = 'FFA500', fill_type = 'solid')
#Yellow
fill_style2 = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')
#Red
fill_style3 = PatternFill(start_color = 'FF8282', end_color = 'FF8282', fill_type = 'solid')


# In[10]:


#Loop through cell
for row in sheet1.iter_rows():
    for cell in row:
        current_cell_value = cell.value
        cell_location = cell.coordinate
        
#if mismatch print old value -> new value       
        if current_cell_value != sheet2[cell_location].value:
            cell.fill = fill_style
            cell.value = f"{current_cell_value} \n -> \n {sheet2[cell_location].value}"
            #Hightlight removed cell
            if current_cell_value is None:
                cell.fill = fill_style2
            #Hightlight new cell
            if sheet2[cell_location].value is None:
                cell.fill = fill_style3
            
#export file 
file1.save("Compared_file.xlsx")

